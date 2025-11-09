#!/usr/bin/env python3
"""
Buyer Data Import Script for Splash25 Migration

This script:
1. Reads buyer data from Excel file
2. Creates user accounts with secure passwords
3. Populates buyer-related tables in the database
4. Exports user credentials to a separate Excel file
"""

import os
import sys
import logging
import random
import string
import subprocess
import pandas as pd
import bcrypt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Check for required libraries
try:
    import psycopg2
    from psycopg2 import sql
except ImportError:
    print("Error: Required Python libraries not found.")
    print("Please install required packages using: pip install -r requirements.txt")
    sys.exit(1)

# Configuration variables
DB_USER = "splash25user"
DB_PASSWORD = "splash25password"
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "splash25_core_db"
EXCEL_FILE = "Demo_buyer.xlsx"
CREDENTIALS_FILE = "buyer_credentials_demo.xlsx"

# Set up logging
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(SCRIPT_DIR, "buyer_import_demo.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Column mappings section
COLUMN_MAPPINGS = {
    "users": {
        "username": "Email",  # Use email as username
        "email": "Email",
        "password_hash": None,  # Will be generated
        "role": None,  # Will be set to 'buyer'
        "business_name": "Company",
        "business_description": None  # Not available in Excel
    },
    "buyer_profiles": {
        "user_id": None,  # Will be set from the created user
        "name": "Name",
        "designation": "Designation",
        "organization": "Company",  # Changed from company to organization to match DB schema
        "mobile": "Mobile",
        "address": "Address",
        "city": "City",
        "state": "State",
        "website": "Website",
        "instagram": "Instagram",
        "operator_type": "Operator Type",
        "country": None,  # Will be set to 'India'
    },
    "buyer_business_info": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "start_year": "Started On",  # Changed from started_on to start_year
        "sell_wayanad": "Promoting Wayanad",  # Using correct field name sell_wayanad
        "sell_wayanad_year": "Promoting Since"  # Using correct field name sell_wayanad_year
    },
    "buyer_demographic": {
        "buyer_id": None,  # Will be set from the created profile
        "mobile": "Mobile",
        "address": "Address",
        "city": "City",
        "state": "State",
        "country": None,  # Will be set to 'India'
        "instagram": "Instagram",
        "company_name": "Company",
        "website": "Website"
        # Note: pincode and gst are not available in Excel
    },
    "buyer_financial_info": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "deposit_paid": None,  # Will be set to False
       # "total_amt_due": None,  # Not available in Excel
       # "total_amt_paid": None  # Not available in Excel
    },
    "buyer_profile_interests": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "interests": "Interests",
        "interest_segment": "Interest Segment"
    },
    "buyer_references": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "ref1_name": "Promoting Property 1",
        "ref1_address": "Promoting Property Address 1",
        "ref2_name": "Promoting Property 2",
        "ref2_address": "Promoting Property Address 2"
    }
}

def check_setup_script_executed():
    """Ask user to confirm setup script has been executed"""
    response = input("Have you already executed setup_migration_db.py? (Yes/No): ")
    if response.lower() != "yes":
        logger.info("Running setup_migration_db.py first...")
        try:
            setup_script = os.path.join(SCRIPT_DIR, "setup_migration_db.py")
            subprocess.run(["python3", setup_script], check=True)
            logger.info("Setup script completed successfully.")
        except subprocess.CalledProcessError as e:
            logger.error(f"Failed to run setup script: {e}")
            sys.exit(1)
    else:
        logger.info("Proceeding with buyer data import...")

def generate_password(length=14):
    """Generate a secure random password"""
    # Define character sets
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    special = "!@#$%^&*()-_=+[]{}|;:,.<>?"
    
    # Ensure at least one character from each set
    password = [
        random.choice(lowercase),
        random.choice(uppercase),
        random.choice(digits),
        random.choice(special)
    ]
    
    # Fill the rest of the password
    remaining_length = length - len(password)
    all_chars = lowercase + uppercase + digits + special
    password.extend(random.choice(all_chars) for _ in range(remaining_length))
    
    # Shuffle the password characters
    random.shuffle(password)
    
    return ''.join(password)

def hash_password(password):
    """Create a bcrypt hash of the password"""
    # Generate a salt and hash the password
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def read_excel_data(file_path):
    """Read data from Excel file"""
    try:
        excel_path = os.path.join(SCRIPT_DIR, file_path)
        df = pd.read_excel(excel_path)
        logger.info(f"Successfully read {len(df)} records from {file_path}")
        return df
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        sys.exit(1)

def get_table_columns(cursor, table_name):
    """Get column names for a database table"""
    try:
        cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}' ORDER BY ordinal_position")
        columns = [row[0] for row in cursor.fetchall()]
        return columns
    except Exception as e:
        logger.error(f"Error getting columns for table {table_name}: {e}")
        return []

def insert_user_data(cursor, user_data):
    """Insert data into users table"""
    try:
        # Generate a password and hash it
        clear_password = generate_password()
        password_hash = hash_password(clear_password)
        
        # Prepare the SQL query
        columns = ["username", "email", "password_hash", "role", "business_name"]
        values = [
            user_data.get("username"),
            user_data.get("email"),
            password_hash,
            "buyer",
            user_data.get("business_name")
        ]
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO users ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        user_id = cursor.fetchone()[0]
        
        logger.info(f"Created user with ID {user_id} for {user_data.get('email')}")
        
        # Return user ID and clear text password
        return user_id, clear_password
    except Exception as e:
        logger.error(f"Error inserting user data: {e}")
        return None, None

def insert_buyer_profile_data(cursor, profile_data, user_id):
    """Insert data into buyer_profiles table"""
    try:
        # Add user_id to profile data
        profile_data["user_id"] = user_id
        
        # Set default values for missing fields
        profile_data["status"] = "active"
        profile_data["country"] = "India"
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in profile_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_profiles ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        profile_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer profile with ID {profile_id} for user {user_id}")
        
        return profile_id
    except Exception as e:
        logger.error(f"Error inserting buyer profile data: {e}")
        return None

def insert_buyer_business_info(cursor, business_data, profile_id):
    """Insert data into buyer_business_info table"""
    try:
        # Add profile_id to business data
        business_data["buyer_profile_id"] = profile_id
        
        # Convert 'True'/'False' string to boolean for sell_wayanad
        if isinstance(business_data.get("sell_wayanad"), str):
            business_data["sell_wayanad"] = business_data.get("sell_wayanad", "False").lower() == "true"
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in business_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_business_info ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        business_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer business info with ID {business_id} for profile {profile_id}")
        
        return business_id
    except Exception as e:
        logger.error(f"Error inserting buyer business info: {e}")
        return None

def insert_buyer_demographic(cursor, demographic_data, profile_id):
    """Insert data into buyer_demographic table"""
    try:
        # Add profile_id to demographic data as buyer_id
        demographic_data["buyer_id"] = profile_id
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in demographic_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        # The buyer_demographic table doesn't have an id column, so we don't use RETURNING id
        query = f"INSERT INTO buyer_demographic ({column_str}) VALUES ({placeholders})"
        
        cursor.execute(query, values)
        
        logger.info(f"Created buyer demographic for profile {profile_id}")
        
        return True
    except Exception as e:
        logger.error(f"Error inserting buyer demographic: {e}")
        return None

def insert_buyer_financial_info(cursor, financial_data, profile_id):
    """Insert data into buyer_financial_info table"""
    try:
        # Add profile_id to financial data
        financial_data["buyer_profile_id"] = profile_id
        
        # Set default values for missing fields
        financial_data["deposit_paid"] = False
        #financial_data["total_amt_due"] = 0
        #financial_data["total_amt_paid"] = 0
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in financial_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_financial_info ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        financial_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer financial info with ID {financial_id} for profile {profile_id}")
        
        return financial_id
    except Exception as e:
        logger.error(f"Error inserting buyer financial info: {e}")
        return None

def insert_buyer_profile_interests(cursor, interests_data, profile_id):
    """Insert data into buyer_profile_interests table"""
    try:
        # Add profile_id to interests data
        interests_data["buyer_profile_id"] = profile_id
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in interests_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_profile_interests ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        interests_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer profile interests with ID {interests_id} for profile {profile_id}")
        
        return interests_id
    except Exception as e:
        logger.error(f"Error inserting buyer profile interests: {e}")
        return None

def insert_buyer_references(cursor, reference_data, profile_id):
    """Insert data into buyer_references table"""
    try:
        # Add profile_id to reference data
        reference_data["buyer_profile_id"] = profile_id
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in reference_data.items():
            if value is not None:
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_references ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        reference_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer references with ID {reference_id} for profile {profile_id}")
        
        return reference_id
    except Exception as e:
        logger.error(f"Error inserting buyer references: {e}")
        return None

def export_credentials(credentials_data):
    """Export user credentials to Excel file"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Buyer Credentials"
        
        # Add headers
        headers = ["User ID", "Username", "Name", "Email", "Password"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_num, user in enumerate(credentials_data, 2):
            ws.cell(row=row_num, column=1).value = user["user_id"]
            ws.cell(row=row_num, column=2).value = user["username"]
            ws.cell(row=row_num, column=3).value = user["name"]
            ws.cell(row=row_num, column=4).value = user["email"]
            ws.cell(row=row_num, column=5).value = user["password"]
        
        # Save the workbook
        output_path = os.path.join(SCRIPT_DIR, CREDENTIALS_FILE)
        wb.save(output_path)
        
        logger.info(f"Exported {len(credentials_data)} user credentials to {CREDENTIALS_FILE}")
        return True
    except Exception as e:
        logger.error(f"Error exporting credentials: {e}")
        return False

def main():
    """Main function to orchestrate the data import process"""
    logger.info("Starting buyer data import process")
    
    # Check if setup script has been executed
    check_setup_script_executed()
    
    # Read Excel data
    excel_data = read_excel_data(EXCEL_FILE)
    
    # Connect to the database
    try:
        conn = psycopg2.connect(
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME
        )
        conn.autocommit = False  # Use transactions
        cursor = conn.cursor()
        logger.info(f"Connected to database {DB_NAME}")
        
        # Store credentials for export
        credentials = []
        
        # Process each row in the Excel file
        for index, row in excel_data.iterrows():
            try:
                # Start a transaction
                cursor.execute("BEGIN")
                
                # Prepare user data
                user_data = {
                    "username": row[COLUMN_MAPPINGS["users"]["username"]],
                    "email": row[COLUMN_MAPPINGS["users"]["email"]],
                    "business_name": row[COLUMN_MAPPINGS["users"]["business_name"]]
                }
                
                # Insert user and get user_id and password
                user_id, password = insert_user_data(cursor, user_data)
                
                if user_id:
                    # Store credentials
                    credentials.append({
                        "user_id": user_id,
                        "username": user_data["username"],
                        "name": row["Name"],
                        "email": user_data["email"],
                        "password": password
                    })
                    
                    # Prepare buyer profile data
                    profile_data = {}
                    for db_col, excel_col in COLUMN_MAPPINGS["buyer_profiles"].items():
                        if excel_col and excel_col in row:
                            profile_data[db_col] = row[excel_col]
                    
                    # Insert buyer profile and get profile_id
                    profile_id = insert_buyer_profile_data(cursor, profile_data, user_id)
                    
                    if profile_id:
                        # Prepare business info data
                        business_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_business_info"].items():
                            if excel_col and excel_col in row:
                                business_data[db_col] = row[excel_col]
                        
                        # Insert business info
                        insert_buyer_business_info(cursor, business_data, profile_id)
                        
                        # Prepare demographic data
                        demographic_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_demographic"].items():
                            if excel_col and excel_col in row:
                                demographic_data[db_col] = row[excel_col]
                        
                        # Insert demographic data
                        #insert_buyer_demographic(cursor, demographic_data, profile_id)
                        
                        # Prepare financial info data
                        financial_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_financial_info"].items():
                            if excel_col and excel_col in row:
                                financial_data[db_col] = row[excel_col]
                        
                        # Insert financial info
                        insert_buyer_financial_info(cursor, financial_data, profile_id)
                        
                        # Prepare profile interests data
                        interests_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_profile_interests"].items():
                            if excel_col and excel_col in row:
                                interests_data[db_col] = row[excel_col]
                        
                        # Insert profile interests
                        #insert_buyer_profile_interests(cursor, interests_data, profile_id)
                        
                        # Prepare references data
                        reference_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_references"].items():
                            if excel_col and excel_col in row:
                                reference_data[db_col] = row[excel_col]
                        
                        # Insert references
                        insert_buyer_references(cursor, reference_data, profile_id)
                
                # Commit the transaction immediately after processing each row
                conn.commit()
                logger.info(f"Successfully imported buyer: {row['Name']} ({row['Email']})")
                
            except Exception as e:
                # Rollback in case of error
                conn.rollback()
                logger.error(f"Error importing buyer {row['Name']} ({row['Email']}): {e}")
        
        # Export credentials to Excel
        export_credentials(credentials)
        
        logger.info(f"Buyer data import completed. Imported {len(credentials)} buyers.")
        
    except Exception as e:
        logger.error(f"Database error: {e}")
    finally:
        # Close database connection
        if conn:
            cursor.close()
            conn.close()
            logger.info("Database connection closed")

if __name__ == "__main__":
    main()
