#!/usr/bin/env python3
"""
Buyer Data Import Script for Splash25 Migration

This script:
1. Reads buyer data from Excel file
2. Creates user accounts with secure passwords
3. Populates buyer-related tables in the database
4. Exports user credentials to a separate Excel file

Usage:
    python import_buyers.py [options]

Options:
    -i, --input-file     Input Excel file path (default: Buyers_20250623.xlsx)
    -d, --db-name        Database name (default: splash25_core_db)
    -h, --db-host        Database host (default: localhost)
    -p, --db-port        Database port (default: 5432)
    -u, --db-user        Database user (default: splash25user)
    -pw, --db-password   Database password (default: splash25password)
    -o, --output-file    Output credentials file (default: buyer_credentials_YYYYMMDD_HHMMSS.xlsx)
"""

import os
import sys
import logging
import random
import string
import subprocess
import argparse
import pandas as pd
import bcrypt
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Check for required libraries
try:
    import psycopg2
    from psycopg2 import sql
except ImportError:
    print("Error: Required Python libraries not found.")
    print("Please install required packages using: pip install -r requirements.txt")
    sys.exit(1)

# Default configuration variables
DEFAULT_DB_USER = "splash25user"
DEFAULT_DB_PASSWORD = "splash25password"
DEFAULT_DB_HOST = "localhost"
DEFAULT_DB_PORT = "5432"
DEFAULT_DB_NAME = "splash25_core_db"
DEFAULT_EXCEL_FILE = "Buyers_20250623.xlsx"
# Generate credentials filename with current date and time
DEFAULT_CREDENTIALS_FILE = f"buyer_credentials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# These will be set by parse_arguments() or use defaults
DB_USER = DEFAULT_DB_USER
DB_PASSWORD = DEFAULT_DB_PASSWORD
DB_HOST = DEFAULT_DB_HOST
DB_PORT = DEFAULT_DB_PORT
DB_NAME = DEFAULT_DB_NAME
EXCEL_FILE = DEFAULT_EXCEL_FILE
CREDENTIALS_FILE = DEFAULT_CREDENTIALS_FILE

# Set up logging
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Generate log filename with current date and time
LOG_FILE = os.path.join(SCRIPT_DIR, f"buyer_import_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Column mappings section
COLUMN_MAPPINGS = {
    "users": {
        "username": "buyer id",  # Use buyer id as username
        "email": "Email",
        "password_hash": None,  # Will be generated
        "role": None,  # Will be set to 'buyer'
        "business_name": "Company",
        "business_description": None  # Not available in Excel
    },
    "buyer_profiles": {
        "user_id": None,  # Will be set from the created user
        "name": "Name",  # Will be parsed into first_name and last_name
        "first_name": None,  # Will be set from parsed Name
        "last_name": None,  # Will be set from parsed Name
        "designation": "Designation",
        "organization": "Company",
        "mobile": "Mobile",
        "address": "Address",
        "city": "City",
        "state": "State",
        "pincode": "Pincode",
        "website": "Website",
        "instagram": "Instagram",
        "gst": "GSTIN",
        "operator_type": "Operator Type",
        "country": None,  # Will be set to 'India'
        "interests": "Interests",  # Will be converted to JSONB
        "properties_of_interest": "Interest Segment",  # Will be converted to JSONB
        "year_of_starting_business": "Started On",
        "selling_wayanad": None,  # Will be set to False
        "vip": None,  # Will be set to False
        "category_id": None,  # Will be set based on buyer_categories where name is "Hosted"
        "since_when": "Promoting Since"
    },
    "buyer_misc": {
        "buyer_id": None,  # Will be set from the created profile
        "previous_visit": "Previous opinion",  # Will be processed with logic
        "previous_stay_property": "Previous stayed property",
        "why_visit": "Reason for Splash Visit"
    },
    "buyer_references": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "ref1_name": "Promoting Property 1",
        "ref1_address": "Promoting Property Address 1",
        "ref2_name": "Promoting Property 2",
        "ref2_address": "Promoting Property Address 2"
    },
    "buyer_financial_info": {
        "buyer_profile_id": None,  # Will be set from the created profile
        "payment_date": "UTR Date",
        "deposit_paid": None,  # Will be set to True
        "entry_fee_paid": None,  # Will be set to True
        "deposit_amount": None,  # Will be set to 0
        "entry_fee_amount": None  # Will be set to 0
    }
}

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description="Import buyer data from Excel to database")
    
    # Add arguments
    parser.add_argument("-i", "--input-file", 
                        help=f"Input Excel file path (default: {DEFAULT_EXCEL_FILE})",
                        default=DEFAULT_EXCEL_FILE)
    
    parser.add_argument("-d", "--db-name", 
                        help=f"Database name (default: {DEFAULT_DB_NAME})",
                        default=DEFAULT_DB_NAME)
    
    parser.add_argument("-H", "--db-host", 
                        help=f"Database host (default: {DEFAULT_DB_HOST})",
                        default=DEFAULT_DB_HOST)
    
    parser.add_argument("-p", "--db-port", 
                        help=f"Database port (default: {DEFAULT_DB_PORT})",
                        default=DEFAULT_DB_PORT)
    
    parser.add_argument("-u", "--db-user", 
                        help=f"Database user (default: {DEFAULT_DB_USER})",
                        default=DEFAULT_DB_USER)
    
    parser.add_argument("-pw", "--db-password", 
                        help=f"Database password (default: {DEFAULT_DB_PASSWORD})",
                        default=DEFAULT_DB_PASSWORD)
    
    parser.add_argument("-o", "--output-file", 
                        help=f"Output credentials file (default: {DEFAULT_CREDENTIALS_FILE})",
                        default=DEFAULT_CREDENTIALS_FILE)
    
    return parser.parse_args()

def check_setup_script_executed():
    """Ask user to confirm setup script has been executed"""
    response = input("Have you already executed setup_migration_db.py? (Yes/No): ")
    if response.lower() != "yes":
        logger.info("Running setup_migration_db.py first...")
        try:
            setup_script = os.path.join(SCRIPT_DIR, "setup_migration_db.py")
            # Pass database connection parameters to setup script
            subprocess.run([
                "python3", 
                setup_script,
                "-d", DB_NAME,
                "-H", DB_HOST,
                "-p", DB_PORT,
                "-u", DB_USER,
                "-pw", DB_PASSWORD
            ], check=True)
            logger.info("Setup script completed successfully.")
        except subprocess.CalledProcessError as e:
            logger.error(f"Failed to run setup script: {e}")
            sys.exit(1)
    else:
        logger.info("Proceeding with buyer data import...")

def generate_password():
    """Generate a secure random password between 8-10 characters"""
    # Define character sets
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    special = "!@#$%^&*()-_=+[]{}|;:,.<>?"
    
    # Randomly choose a length between 8 and 10
    length = random.randint(8, 10)
    
    # Ensure at least one character from each set
    password = [
        random.choice(lowercase),
        random.choice(uppercase),
        random.choice(digits),
        random.choice(special)
    ]
    
    # Fill the rest of the password
    remaining_length = length - len(password)
    all_chars = lowercase + uppercase + digits + special
    password.extend(random.choice(all_chars) for _ in range(remaining_length))
    
    # Shuffle the password characters
    random.shuffle(password)
    
    return ''.join(password)

def hash_password(password):
    """Create a bcrypt hash of the password using BF salt"""
    # Generate a salt and hash the password
    salt = bcrypt.gensalt()  # Uses BF by default
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def parse_name(full_name):
    """Parse full name into first name and last name with proper capitalization"""
    if not full_name or pd.isna(full_name):
        return "", ""
    
    # Split the name by spaces
    name_parts = full_name.strip().split()
    
    if len(name_parts) == 0:
        return "", ""
    elif len(name_parts) == 1:
        # Only one name provided, treat as first name
        return name_parts[0].capitalize(), ""
    else:
        # First part is first name, rest is last name
        first_name = name_parts[0].capitalize()
        last_name = " ".join(part.capitalize() for part in name_parts[1:])
        return first_name, last_name

def convert_to_jsonb(value):
    """Convert a string value to JSONB format"""
    if not value or pd.isna(value):
        return json.dumps([])
    
    # Split by commas and clean up each item
    items = [item.strip() for item in str(value).split(',') if item.strip()]
    
    # Return as a JSON array string
    return json.dumps(items)

def get_hosted_category_id(cursor):
    """Get the category ID for 'Hosted' from buyer_categories table"""
    try:
        cursor.execute("SELECT id FROM buyer_categories WHERE name = 'Hosted'")
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            logger.warning("Category 'Hosted' not found in buyer_categories table, using default value 1")
            return 1
    except Exception as e:
        logger.error(f"Error getting Hosted category ID: {e}")
        return 1

def read_excel_data(file_path):
    """Read data from Excel file"""
    try:
        excel_path = os.path.join(SCRIPT_DIR, file_path)
        df = pd.read_excel(excel_path)
        logger.info(f"Successfully read {len(df)} records from {file_path}")
        return df
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        sys.exit(1)

def get_table_columns(cursor, table_name):
    """Get column names for a database table"""
    try:
        cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}' ORDER BY ordinal_position")
        columns = [row[0] for row in cursor.fetchall()]
        return columns
    except Exception as e:
        logger.error(f"Error getting columns for table {table_name}: {e}")
        return []

def insert_user_data(cursor, user_data):
    """Insert data into users table"""
    try:
        # Generate a password and hash it
        clear_password = generate_password()
        password_hash = hash_password(clear_password)
        
        # Prepare the SQL query
        columns = ["username", "email", "password_hash", "role", "business_name"]
        values = [
            user_data.get("username"),
            user_data.get("email"),
            password_hash,
            "buyer",
            user_data.get("business_name")
        ]
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO users ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        user_id = cursor.fetchone()[0]
        
        logger.info(f"Created user with ID {user_id} for {user_data.get('email')}")
        
        # Return user ID and clear text password
        return user_id, clear_password
    except Exception as e:
        logger.error(f"Error inserting user data: {e}")
        return None, None

def insert_buyer_profile_data(cursor, profile_data, user_id):
    """Insert data into buyer_profiles table"""
    try:
        # Add user_id to profile data
        profile_data["user_id"] = user_id
        
        # Set default values for missing fields
        profile_data["status"] = "active"
        profile_data["country"] = "India"
        profile_data["selling_wayanad"] = False
        profile_data["vip"] = False
        
        # Get the category_id for "Hosted"
        profile_data["category_id"] = get_hosted_category_id(cursor)
        
        # Parse name into first_name and last_name
        if "name" in profile_data:
            first_name, last_name = parse_name(profile_data["name"])
            profile_data["first_name"] = first_name
            profile_data["last_name"] = last_name
        
        # Convert interests and properties_of_interest to JSONB
        if "interests" in profile_data and profile_data["interests"] is not None:
            profile_data["interests"] = convert_to_jsonb(profile_data["interests"])
        
        if "properties_of_interest" in profile_data and profile_data["properties_of_interest"] is not None:
            profile_data["properties_of_interest"] = convert_to_jsonb(profile_data["properties_of_interest"])
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in profile_data.items():
            if value is not None and key in get_table_columns(cursor, "buyer_profiles"):
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_profiles ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        profile_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer profile with ID {profile_id} for user {user_id}")
        
        return profile_id
    except Exception as e:
        logger.error(f"Error inserting buyer profile data: {e}")
        return None

def insert_buyer_misc(cursor, misc_data, profile_id):
    """Insert data into buyer_misc table"""
    try:
        # Add profile_id to misc data as buyer_id
        misc_data["buyer_id"] = profile_id
        
        # Process previous_visit based on "Previous opinion" field
        if "previous_visit" in misc_data:
            previous_opinion = misc_data["previous_visit"]
            if pd.isna(previous_opinion) or previous_opinion == "I haven't attended before":
                misc_data["previous_visit"] = False
            else:
                misc_data["previous_visit"] = True
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in misc_data.items():
            if value is not None and key in get_table_columns(cursor, "buyer_misc"):
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_misc ({column_str}) VALUES ({placeholders})"
        
        cursor.execute(query, values)
        
        logger.info(f"Created buyer misc data for profile {profile_id}")
        
        return True
    except Exception as e:
        logger.error(f"Error inserting buyer misc data: {e}")
        return None

def insert_buyer_references(cursor, reference_data, profile_id):
    """Insert data into buyer_references table"""
    try:
        # Add profile_id to reference data
        reference_data["buyer_profile_id"] = profile_id
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in reference_data.items():
            if value is not None and key in get_table_columns(cursor, "buyer_references"):
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_references ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        reference_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer references with ID {reference_id} for profile {profile_id}")
        
        return reference_id
    except Exception as e:
        logger.error(f"Error inserting buyer references: {e}")
        return None

def insert_buyer_financial_info(cursor, financial_data, profile_id):
    """Insert data into buyer_financial_info table"""
    try:
        # Add profile_id to financial data
        financial_data["buyer_profile_id"] = profile_id
        
        # Set default values for missing fields
        financial_data["deposit_paid"] = True
        financial_data["entry_fee_paid"] = True
        financial_data["deposit_amount"] = 0
        financial_data["entry_fee_amount"] = 0
        
        # Prepare the SQL query
        columns = []
        values = []
        
        for key, value in financial_data.items():
            if value is not None and key in get_table_columns(cursor, "buyer_financial_info"):
                columns.append(key)
                values.append(value)
        
        placeholders = ", ".join(["%s"] * len(values))
        column_str = ", ".join(columns)
        
        query = f"INSERT INTO buyer_financial_info ({column_str}) VALUES ({placeholders}) RETURNING id"
        
        cursor.execute(query, values)
        financial_id = cursor.fetchone()[0]
        
        logger.info(f"Created buyer financial info with ID {financial_id} for profile {profile_id}")
        
        return financial_id
    except Exception as e:
        logger.error(f"Error inserting buyer financial info: {e}")
        return None

def export_credentials(credentials_data):
    """Export user credentials to Excel file"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = "Buyer Credentials"
            
            # Add headers
            headers = ["User ID", "First Name", "Last Name", "Email", "Mobile", "Username", "Password"]
            for col_num, header in enumerate(headers, 1):
                try:
                    ws.cell(row=1, column=col_num).value = header
                    ws.cell(row=1, column=col_num).font = Font(bold=True)
                    ws.cell(row=1, column=col_num).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                except AttributeError:
                    # Handle potential MergedCell issues
                    logger.warning(f"Could not format header cell at column {col_num}")
            
            # Add data
            for row_num, user in enumerate(credentials_data, 2):
                try:
                    # Set values directly
                    ws.cell(row=row_num, column=1).value = user["user_id"]
                    ws.cell(row=row_num, column=2).value = user["first_name"]
                    ws.cell(row=row_num, column=3).value = user["last_name"]
                    ws.cell(row=row_num, column=4).value = user["email"]
                    ws.cell(row=row_num, column=5).value = user["mobile"]
                    ws.cell(row=row_num, column=6).value = user["username"]
                    ws.cell(row=row_num, column=7).value = user["password"]
                except AttributeError:
                    # Handle potential MergedCell issues
                    logger.warning(f"Could not set values for row {row_num}")
        
        # Save the workbook
        output_path = os.path.join(SCRIPT_DIR, CREDENTIALS_FILE)
        wb.save(output_path)
        
        logger.info(f"Exported {len(credentials_data)} user credentials to {CREDENTIALS_FILE}")
        return True
    except Exception as e:
        logger.error(f"Error exporting credentials: {e}")
        return False

def main():
    """Main function to orchestrate the data import process"""
    # Parse command line arguments
    global DB_USER, DB_PASSWORD, DB_HOST, DB_PORT, DB_NAME, EXCEL_FILE, CREDENTIALS_FILE
    args = parse_arguments()
    
    # Set configuration variables from arguments
    DB_USER = args.db_user
    DB_PASSWORD = args.db_password
    DB_HOST = args.db_host
    DB_PORT = args.db_port
    DB_NAME = args.db_name
    EXCEL_FILE = args.input_file
    CREDENTIALS_FILE = args.output_file
    
    logger.info("Starting buyer data import process")
    logger.info(f"Using Excel file: {EXCEL_FILE}")
    logger.info(f"Using database: {DB_NAME} on {DB_HOST}:{DB_PORT}")
    logger.info(f"Output credentials will be saved to: {CREDENTIALS_FILE}")
    
    # Check if setup script has been executed
    check_setup_script_executed()
    
    # Read Excel data
    excel_data = read_excel_data(EXCEL_FILE)
    
    # Connect to the database
    try:
        conn = psycopg2.connect(
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME
        )
        conn.autocommit = False  # Use transactions
        cursor = conn.cursor()
        logger.info(f"Connected to database {DB_NAME}")
        
        # Store credentials for export
        credentials = []
        
        # Process each row in the Excel file
        for index, row in excel_data.iterrows():
            try:
                # Start a transaction
                cursor.execute("BEGIN")
                
                # Prepare user data
                user_data = {
                    "username": row[COLUMN_MAPPINGS["users"]["username"]],
                    "email": row[COLUMN_MAPPINGS["users"]["email"]],
                    "business_name": row[COLUMN_MAPPINGS["users"]["business_name"]]
                }
                
                # Insert user and get user_id and password
                user_id, password = insert_user_data(cursor, user_data)
                
                if user_id:
                    # Parse name for credentials
                    first_name, last_name = parse_name(row["Name"])
                    
                    # Store credentials
                    credentials.append({
                        "user_id": user_id,
                        "username": user_data["username"],
                        "first_name": first_name,
                        "last_name": last_name,
                        "email": user_data["email"],
                        "mobile": row["Mobile"] if "Mobile" in row else "",
                        "password": password
                    })
                    
                    # Prepare buyer profile data
                    profile_data = {}
                    for db_col, excel_col in COLUMN_MAPPINGS["buyer_profiles"].items():
                        if excel_col and excel_col in row:
                            profile_data[db_col] = row[excel_col]
                    
                    # Insert buyer profile and get profile_id
                    profile_id = insert_buyer_profile_data(cursor, profile_data, user_id)
                    
                    if profile_id:
                        # Prepare references data
                        reference_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_references"].items():
                            if excel_col and excel_col in row:
                                reference_data[db_col] = row[excel_col]
                        
                        # Insert references
                        insert_buyer_references(cursor, reference_data, profile_id)
                        
                        # Prepare misc data
                        misc_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_misc"].items():
                            if excel_col and excel_col in row:
                                misc_data[db_col] = row[excel_col]
                        
                        # Insert misc data
                        insert_buyer_misc(cursor, misc_data, profile_id)
                        
                        # Prepare financial info data
                        financial_data = {}
                        for db_col, excel_col in COLUMN_MAPPINGS["buyer_financial_info"].items():
                            if excel_col and excel_col in row:
                                financial_data[db_col] = row[excel_col]
                        
                        # Insert financial info
                        insert_buyer_financial_info(cursor, financial_data, profile_id)
                
                # Commit the transaction immediately after processing each row
                conn.commit()
                logger.info(f"Successfully imported buyer: {row['Name']} ({row['Email']})")
                
            except Exception as e:
                # Rollback in case of error
                conn.rollback()
                logger.error(f"Error importing buyer {row['Name']} ({row['Email']}): {e}")
        
        # Export credentials to Excel
        export_credentials(credentials)
        
        logger.info(f"Buyer data import completed. Imported {len(credentials)} buyers.")
        
    except Exception as e:
        logger.error(f"Database error: {e}")
    finally:
        # Close database connection
        if conn:
            cursor.close()
            conn.close()
            logger.info("Database connection closed")

if __name__ == "__main__":
    main()
